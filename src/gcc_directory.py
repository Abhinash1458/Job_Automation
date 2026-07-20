"""GCC company directory — scraped from businessofgcc.com (Bangalore + Hyderabad).

Used as a COMPANY ALLOWLIST so the daily pipeline keeps only jobs at real GCCs /
product captive centers and drops IT-services firms. Refreshed daily and cached
in the `gccs` table so a scrape failure never empties the filter.
"""
from __future__ import annotations

import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timezone

import requests

from . import config

CITIES = ("bangalore", "hyderabad")
_BASE = "https://www.businessofgcc.com/gcc-data/companies/cities/{city}"
_H2_RE = re.compile(r'<h2[^>]*text-\[#101828\][^>]*>([^<]+)</h2>')
_HEADERS = {"User-Agent": "Mozilla/5.0 (job-hunt-automation research)"}

# Generic words stripped when deriving a brand keyword from a GCC's full name.
_GENERIC = {
    "india", "indian", "global", "capability", "center", "centre", "centers",
    "centres", "technology", "technologies", "engineering", "solutions",
    "services", "business", "shared", "software", "innovation", "operations",
    "development", "delivery", "network", "labs", "lab", "gcc", "of", "and",
    "the", "systems", "group", "digital", "tech", "ltd", "pvt", "private",
    "limited", "co", "company", "international", "worldwide", "east", "west",
    "north", "south", "&", "healthcare", "health", "industries", "financial",
    "technical", "connected", "design", "acceleration", "hub", "hubs", "center's",
}

# Seed brands so the allowlist works even before/without a successful scrape.
# (Apple / Google / Goldman / Atlassian intentionally excluded per user request;
#  PwC & KPMG intentionally included per user request.)
_SEED = [
    "Microsoft", "Amazon", "Wells Fargo",
    "JPMorgan", "Morgan Stanley", "Adobe", "Salesforce", "ServiceNow",
    "Uber", "Walmart", "Bloomberg", "AMD", "Nvidia", "Qualcomm", "Cisco", "Dell",
    "Intuit", "Visa", "Mastercard", "American Express", "Target",
    "Lowe's", "Optum", "UnitedHealth", "Fidelity", "BlackRock", "Cloudflare",
    "Dropbox", "Booking.com", "Expedia", "LinkedIn", "SAP", "VMware",
    "Broadcom", "Arm", "Analog Devices", "Micron", "Texas Instruments", "Airbus",
    "Boeing", "General Motors", "Bosch", "Continental", "AstraZeneca", "Novartis",
    "Pfizer", "Eli Lilly", "Bayer", "GE HealthCare", "Cargill", "ExxonMobil",
    "Shell", "Equifax", "Cigna", "Evernorth", "Autodesk", "Cadence", "Synopsys",
    "Applied Materials", "IBM", "HP", "HPE", "Sprinklr", "Freshworks", "PhonePe",
    "Razorpay", "Swiggy", "Meesho", "Zomato", "CRED", "Postman",
    "PwC", "KPMG",  # user explicitly wants these consulting GCCs included
]

# Companies to always DROP even if they appear in the scraped directory / a job's
# company field matches something. User-specified exclusions.
_EXCLUDE_BRANDS = {
    "apple", "google", "goldman", "goldman sachs", "atlassian",
    "flipkart", "paypal", "oracle",
}


@contextmanager
def _conn():
    config.DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute(
            "CREATE TABLE IF NOT EXISTS gccs ("
            "name TEXT PRIMARY KEY, brand TEXT, city TEXT, updated TEXT)"
        )
        yield conn
        conn.commit()
    finally:
        conn.close()


def _brand(name: str) -> str:
    """Reduce a full GCC name to a lowercase brand keyword for matching."""
    cleaned = name.replace("&amp;", "&")
    tokens = re.split(r"[\s/,]+", cleaned)
    kept = []
    for t in tokens:
        tok = t.strip(".&()[]-'").strip()          # drop wrapping punctuation
        if tok and tok.lower() not in _GENERIC:
            kept.append(tok.lower())
    # keep at most the first 2 brand tokens (e.g. "goldman sachs", "wells fargo")
    return " ".join(kept[:2]) if kept else name.lower()


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def scrape_city(city: str) -> list[str]:
    resp = requests.get(_BASE.format(city=city), headers=_HEADERS, timeout=30)
    resp.raise_for_status()
    names = {m.group(1).replace("&amp;", "&").strip() for m in _H2_RE.finditer(resp.text)}
    return sorted(names)


_TD_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL)
_CELL_RE = re.compile(r"<td[^>]*>(.*?)</td>", re.DOTALL)


def scrape_gccjournal() -> list[str]:
    """Parse the GCC company table on gccjournal.in (first column = company)."""
    import html as _html

    url = "https://gccjournal.in/insights/list-of-global-capability-centers-gcc-in-india/"
    resp = requests.get(url, headers=_HEADERS, timeout=30)
    resp.raise_for_status()
    names = set()
    for row in _TD_RE.finditer(resp.text):
        cells = _CELL_RE.findall(row.group(1))
        if not cells:
            continue
        name = _html.unescape(re.sub(r"<[^>]+>", "", cells[0])).strip()
        if name and len(name) < 50 and not name[0].isdigit():
            names.add(name)
    return sorted(names)


# Curated company allowlist (authoritative). If present, it is the SOLE source
# of truth for is_gcc — web scrapers and the seed list are ignored.
COMPANIES_XLSX = config.DATA_DIR / "companies.xlsx"


def load_excel() -> list[tuple[str, str]]:
    """Return (company, city) rows from data/companies.xlsx, or [] if absent."""
    if not COMPANIES_XLSX.exists():
        return []
    import openpyxl

    wb = openpyxl.load_workbook(COMPANIES_XLSX, read_only=True, data_only=True)
    ws = wb["GCCs & Product Cos"] if "GCCs & Product Cos" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = next((i for i, r in enumerate(rows) if r and r[0] == "Company"), None)
    if hdr is None:
        return []
    out = []
    for r in rows[hdr + 1:]:
        if r and r[0]:
            city = str(r[1]).strip() if len(r) > 1 and r[1] else "India"
            out.append((str(r[0]).strip(), city))
    return out


def refresh(force: bool = False) -> int:
    """Rebuild the gccs allowlist. Prefers data/companies.xlsx; otherwise falls
    back to the businessofgcc / gccjournal scrapers + seed list."""
    with _conn() as c:
        if not force:
            row = c.execute("SELECT MAX(updated) u FROM gccs").fetchone()
            if row and row["u"] and row["u"][:10] == _now()[:10]:
                return c.execute("SELECT COUNT(*) n FROM gccs").fetchone()["n"]

    scraped: dict[str, str] = {}
    excel = load_excel()
    if excel:
        # Excel is authoritative — use it exclusively.
        for name, city in excel:
            scraped[name] = f"excel:{city}"
    else:
        for city in CITIES:
            try:
                for name in scrape_city(city):
                    scraped[name] = city
            except Exception as exc:  # noqa: BLE001
                print(f"  GCC scrape failed for {city}: {exc}")
        try:
            for name in scrape_gccjournal():
                scraped.setdefault(name, "gccjournal")
        except Exception as exc:  # noqa: BLE001
            print(f"  gccjournal scrape failed: {exc}")
        for name in _SEED:
            scraped.setdefault(name, "seed")

    # drop any user-excluded companies from the stored directory
    scraped = {n: c for n, c in scraped.items()
               if not any(x in n.lower() for x in _EXCLUDE_BRANDS)}

    with _conn() as c:
        c.execute("DELETE FROM gccs")  # rebuild cleanly so removed companies drop out
        for name, city in scraped.items():
            c.execute(
                "INSERT INTO gccs (name, brand, city, updated) VALUES (?,?,?,?) "
                "ON CONFLICT(name) DO UPDATE SET brand=excluded.brand, "
                "city=excluded.city, updated=excluded.updated",
                (name, _brand(name), city, _now()),
            )
        return c.execute("SELECT COUNT(*) n FROM gccs").fetchone()["n"]


def brands() -> list[str]:
    with _conn() as c:
        rows = c.execute("SELECT DISTINCT brand FROM gccs WHERE brand != ''").fetchall()
        if rows:
            return sorted({r["brand"] for r in rows})
    # fallback only when the table is empty (never refreshed)
    return sorted({_brand(s) for s in _SEED})


# IT-services firms to always exclude, even if the directory lists them (user
# wants product companies / captive GCCs only).
_SERVICES_BLOCKLIST = {
    "cognizant", "infosys", "tcs", "tata consultancy", "wipro", "accenture",
    "capgemini", "hcl", "tech mahindra", "ltimindtree", "mindtree", "mphasis",
    "persistent", "coforge", "birlasoft", "zensar", "hexaware", "cyient",
    "deloitte", "ernst", "genpact", "wns", "nagarro",
    # user-specified exclusions
    "flipkart", "paypal", "oracle",
}


def is_gcc(company: str) -> bool:
    """True if a job's company name matches a known GCC brand (and is not a
    blocklisted IT-services firm or a user-excluded company)."""
    if not company:
        return False
    comp = company.lower()
    if any(b in comp for b in _SERVICES_BLOCKLIST):
        return False
    if any(b in comp for b in _EXCLUDE_BRANDS):
        return False
    comp_words = set(re.findall(r"[a-z0-9]+", comp))
    for b in brands():
        bwords = b.split()
        if not bwords:
            continue
        if len(bwords) == 1:
            w = bwords[0]
            # distinctive single-word brand: whole-word match; very short: substring
            if (len(w) >= 3 and w in comp_words) or (len(w) < 3 and w in comp):
                return True
        else:
            # multi-word brand: require all tokens present (avoids "american" etc.)
            if all(w in comp_words for w in bwords) or b in comp:
                return True
    return False


def count() -> int:
    with _conn() as c:
        return c.execute("SELECT COUNT(*) n FROM gccs").fetchone()["n"]
